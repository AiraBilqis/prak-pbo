####################################  STUDENT REGISTRATION  ####################################
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#06285B"
framebg = "#EDEDED"
framefg = "#06285B"

root = Tk()
root.title("Student Registration System")
root.geometry("1240x700+220+100")
root.config(bg = background)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Date of Birth"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Study Program"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    
    file.save('Student_data.xlsx')

#Button Exit Window Function
def Exit():
    root.destroy()

#Button Upload (Show Image) Function
def showimage():
    global filename
    global img
    
    filename = filedialog.askopenfilename(initialdir = os.getcwd(),
                                          title = "Select Image File", filetype = (("JPG File", "*.jpg"),
                                                                                   ("PNG File", "*.png"),
                                                                                   ("All files", "*.txt")))
    
    img = (Image.open(filename))
    resized_image = img.resize((197, 197))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image = photo2)
    lbl.image = photo2

#Registration No.
def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    
    max_row_value = sheet.cell(row = row, column = 1).value
    
    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")

#Button Clear Function
def Clear():
    global img
    
    Name.set('')
    DOB.set('')
    Religion.set('')
    Study.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")
    
    registration_no()
    
    saveButton.config(state = 'normal')
    
    img1 = PhotoImage(file = 'Student_Images/Upload Photo.png')
    lbl.config(image = img1)
    lbl.image = img1
    
    img = ""
    
#Button Save Function
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender #Jika tidak memilih gender akan muncul pesan error
    except:
        messagebox.showerror("Error", "Select Gender!")
    
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Study.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()
    
    if N1 == "" or C1 == "Select Class" or D2 == "" or Re1 == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("Error", "Few Data is Mising!")
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column = 1, row = sheet.max_row + 1, value = R1)
        sheet.cell(column = 2, row = sheet.max_row, value = N1)
        sheet.cell(column = 3, row = sheet.max_row, value = C1)
        sheet.cell(column = 4, row = sheet.max_row, value = G1)
        sheet.cell(column = 5, row = sheet.max_row, value = D2)
        sheet.cell(column = 6, row = sheet.max_row, value = D1)
        sheet.cell(column = 7, row = sheet.max_row, value = Re1)
        sheet.cell(column = 8, row = sheet.max_row, value = S1)
        sheet.cell(column = 9, row = sheet.max_row, value = fathername)
        sheet.cell(column = 10, row = sheet.max_row, value = mothername)
        sheet.cell(column = 11, row = sheet.max_row, value = F1)
        sheet.cell(column = 12, row = sheet.max_row, value = M1)
        
        file.save(r'Student_data.xlsx')
        
        try:
            img.save("Student Images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("Info", "Profile Picture is Not Available!!!")
        
        messagebox.showinfo("Info", "Sucessfully Data Entered!!!")
        
        Clear() #Clear entry box and image section
        
        registration_no() #It will recheck registration no. and reissue new no.

#Button Search Function
def search():
    text = Search.get() #Taking input from entry box
    
    Clear() #To clear all the data already available in entry box and other
    saveButton.config(state = 'disabled')
    
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active
    
    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            #print(str(name))
            ref_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            
            #print(reg_no_position)
            #print(reg_number)
            
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid Registration Number!!!")
    
    #reg_no_position showing like A2, A3, A4, ...., An
    #but reg_number just showing number after A2 like 2, 3, ...., n
    
    x1 = sheet.cell(row = int(reg_number), column = 1).value
    x2 = sheet.cell(row = int(reg_number), column = 2).value
    x3 = sheet.cell(row = int(reg_number), column = 3).value
    x4 = sheet.cell(row = int(reg_number), column = 4).value
    x5 = sheet.cell(row = int(reg_number), column = 5).value
    x6 = sheet.cell(row = int(reg_number), column = 6).value
    x7 = sheet.cell(row = int(reg_number), column = 7).value
    x8 = sheet.cell(row = int(reg_number), column = 8).value
    x9 = sheet.cell(row = int(reg_number), column = 9).value
    x10 = sheet.cell(row = int(reg_number), column = 10).value
    x11 = sheet.cell(row = int(reg_number), column = 11).value
    x12 = sheet.cell(row = int(reg_number), column = 12).value
    
    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)
    
    if x4 == 'Female':
        R2.select()
    else:
        R1.select()
        
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Study.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)
    
    img = (Image.open("Student Images/" + str(x1) + ".jpg")) #We done this to take image name same as registration no.
    resized_image = img.resize((197, 197))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image = photo2)
    lbl.image = photo2
    
#Update
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Study.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()
    
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active
    
    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            
            print(reg_number)
    
    # sheet.cell(column = 1, row = int(reg_number), value = R1) now noone can update registration no.
    sheet.cell(column = 2, row = int(reg_number), value = N1)
    sheet.cell(column = 3, row = int(reg_number), value = C1)
    sheet.cell(column = 4, row = int(reg_number), value = G1)
    sheet.cell(column = 5, row = int(reg_number), value = D2)
    sheet.cell(column = 6, row = int(reg_number), value = D1)
    sheet.cell(column = 7, row = int(reg_number), value = Re1)
    sheet.cell(column = 8, row = int(reg_number), value = S1)
    sheet.cell(column = 9, row = int(reg_number), value = fathername)
    sheet.cell(column = 10, row = int(reg_number), value = mothername)
    sheet.cell(column = 11, row = int(reg_number), value = F1)
    sheet.cell(column = 12, row = int(reg_number), value = M1)
    
    file.save(r'Student_data.xlsx')
    
    try:
        img.save("Student Images/" + str(R1) + ".jpg")
    except:
        pass
    
    messagebox.showinfo("Update", "Update Sucessfully!!!")
    
    Clear() #After update it will clear all entry boxes
    
#gender
def selection():
    global gender
    
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

#Top Frames
Label(root, text = "Email : khairani.121140091@student.itera.ac.id", width = 10, height = 3, bg  = "#f0687c", anchor = 'e').pack(side = TOP, fill = X)
Label(root, text = "STUDENT REGISTRATION", width = 10, height = 2, bg  = "#c36464", fg = '#fff', font = 'calibri 20 bold').pack(side = TOP, fill = X)

#Search Box to Update
Search = StringVar()
Entry(root, textvariable = Search, width = 15, bd = 2, font = "calibri 20").place(x = 830, y = 67)
imageicon3 = PhotoImage(file = "Student_Images/Search Icon.png")
Srch = Button(root, text = "Search", compound = LEFT, image = imageicon3, width = 100, bg = '#68ddfa', font = "timesnewroman 13 bold", command = search)
Srch.place(x = 1060, y = 70)

imageicon4 = PhotoImage(file = "Student_Images/Layer.png")
Update_button = Button(root, image = imageicon4, bg = "#c36464", command = Update)
Update_button.place(x = 110, y = 64)

#Registration and Date
Label(root, text = "Registration No:", font = "calibri 13", fg = framebg, bg = background).place(x = 30, y = 150)
Label(root, text = "Date:", font = "calibri 13", fg = framebg, bg = background).place(x = 500, y = 150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable = Registration, width = 25, font = "calibri 10")
reg_entry.place(x = 160, y = 155)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable = Date, width = 15, font = "arial 10")
date_entry.place(x = 550, y = 155)

Date.set(d1)

#Student Details
obj = LabelFrame(root, text = "Student's Details", font = 18, bd = 2, width = 850, bg = framebg, fg = framefg, height = 230, relief = GROOVE)
obj.place(x = 30, y = 200)

Label(obj, text = "Full Name:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 30, y = 40)
Label(obj, text = "Date of Birth:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 30, y = 90)
Label(obj, text = "Gender:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 30, y = 140)

Label(obj, text = "Class:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 440, y = 40)
Label(obj, text = "Religion:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 440, y = 90)
Label(obj, text = "Study Program:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 440, y = 140)

Name = StringVar()
name_entry = Entry(obj, textvariable = Name, width = 30, font = "arial 10")
name_entry.place(x = 145, y = 43)

DOB = StringVar()
dob_entry = Entry(obj, textvariable = DOB, width = 30, font = "arial 10")
dob_entry.place(x = 145, y = 93)

radio = IntVar()
R1 = Radiobutton(obj, text = "Male", variable = radio, value = 1, bg = framebg, fg = framefg, command = selection)
R1.place(x = 145, y = 143)

R2 = Radiobutton(obj, text = "Female", variable = radio, value = 2, bg = framebg, fg = framefg, command = selection)
R2.place(x = 205, y = 143)

Religion = StringVar()
religion_entry = Entry(obj, textvariable = Religion, width = 30, font = "arial 10")
religion_entry.place(x = 580, y = 93)

Study = StringVar()
study_entry = Entry(obj, textvariable = Study, width = 30, font = "arial 10")
study_entry.place(x = 580, y = 143)

Class = Combobox(obj, values = ['RA', 'RB', 'RC', 'RD'], font = "arial 10", width = 27, state = "r")
Class.place(x = 580, y = 43)
Class.set("Select Class")

#Parents Details
obj2 = LabelFrame(root, text = "Parent's Details", font = 18, bd = 2, width = 850, bg = framebg, fg = framefg, height = 200, relief = GROOVE)
obj2.place(x = 30, y = 450)

Label(obj2, text = "Father's Name:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 30, y = 40)
Label(obj2, text = "Occupation:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 30, y = 90)

F_Name = StringVar()
F_entry = Entry(obj2, textvariable = F_Name, width = 30, font = "arial 10")
F_entry.place(x = 145, y = 43)

Father_Occupation = StringVar()
FO_entry = Entry(obj2, textvariable = Father_Occupation, width = 30, font = "arial 10")
FO_entry.place(x = 145, y = 93)

Label(obj2, text = "Mother's Name:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 440, y = 40)
Label(obj2, text = "Occupation:", font = "calibri 13", bg = framebg, fg = framefg).place(x = 440, y = 90)

M_Name = StringVar()
M_entry = Entry(obj2, textvariable = M_Name, width = 30, font = "arial 10")
M_entry.place(x = 580, y = 43)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2, textvariable = Mother_Occupation, width = 30, font = "arial 10")
MO_entry.place(x = 580, y = 93)

#Image
f = Frame(root, bd = 2, bg = "black", width = 207, height = 207, relief = GROOVE)
f.place(x = 950, y = 150)

img = PhotoImage(file = "Student_Images/Upload Photo1.png") 
lbl = Label(f, bg = "black", image = img)
lbl.place(x = 0, y = 0)

#Button Upload Image
Button(root, text = "Upload", width = 20, height = 1, font = "arial 12 bold", bg = "lightblue", command = showimage).place(x = 950, y = 370)

#Button Save
saveButton = Button(root, text = "Save", width = 20, height = 1, font = "arial 12 bold", bg = "lightgreen", command = Save)
saveButton.place(x = 950, y = 430)

#Button Reset
Button(root, text = "Reset", width = 20, height = 1, font = "arial 12 bold", bg = "lightpink", command = Clear).place(x = 950, y = 490)

#Button Exit
Button(root, text = "Exit", width = 20, height = 1, font = "arial 12 bold", bg = "red", command = Exit).place(x = 950, y = 560)


root.mainloop()